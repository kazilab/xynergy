from importlib.resources import as_file, files

import numpy as np
import polars as pl
from polars import col

from xynergy.fit import ll4

EXAMPLE_COLUMN_RENAME_MAP = {
    "Experiment_ID": "experiment_source_id",
    "Bio_specimen": "line",
    "PairIndex": "pair_index",
    "Drug1": "drug_a",
    "Drug2": "drug_b",
    "Conc1": "dose_a",
    "Conc2": "dose_b",
    "Response": "response",
}

EXAMPLE_DOSING_COLUMNS = ["dose_a", "dose_b"]
EXAMPLE_RESPONSE_COLUMN = "response"
EXAMPLE_EXPERIMENT_COLUMNS = [
    "experiment_source_id",
    "line",
    "drug_a",
    "drug_b",
    "pair_index",
]


def load_example_data(
    raw: bool = False,
    convert_to_inhibition: bool = True,
) -> pl.DataFrame:
    """Load the bundled workbook example.

    Parameters
    ----------
    raw : bool, default False
        If `True`, keep the original workbook column names. Otherwise rename
        columns to Xynergy's canonical names.
    convert_to_inhibition : bool, default True
        The workbook stores viability-style percentages where 100 means no
        effect and 0 means complete effect. If `True`, convert this to
        inhibition-style percentages where 0 means no effect and 100 means
        complete effect.

    Returns
    -------
    polars.DataFrame
        The bundled example dataset.
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is required to load the bundled Excel example data."
        ) from exc

    resource = files("xynergy").joinpath("example_data/data.xlsx")
    with as_file(resource) as path:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

    headers = [str(value) for value in rows[0]]
    records = [
        dict(zip(headers, row, strict=False))
        for row in rows[1:]
        if any(value is not None for value in row)
    ]
    df = pl.DataFrame(records)

    if raw:
        response_col = "Response"
        dose_cols = ["Conc1", "Conc2"]
    else:
        df = df.rename(EXAMPLE_COLUMN_RENAME_MAP)
        response_col = EXAMPLE_RESPONSE_COLUMN
        dose_cols = EXAMPLE_DOSING_COLUMNS

    df = df.with_columns(pl.col(response_col).cast(pl.Float64))
    if convert_to_inhibition:
        df = df.with_columns((100.0 - pl.col(response_col)).alias(response_col))

    return df.sort(dose_cols)


def get_example_xynergy_kwargs() -> dict[str, object]:
    """Return defaults for running the bundled workbook through Xynergy."""
    return {
        "dose_cols": EXAMPLE_DOSING_COLUMNS.copy(),
        "response_col": EXAMPLE_RESPONSE_COLUMN,
        "experiment_cols": EXAMPLE_EXPERIMENT_COLUMNS.copy(),
        "response_is_percent": True,
        "complete_response_is_0": False,
    }


# n_doses does not include 0 dose
def make_example_data(
    n_doses=5,
    dose_start=0.01,
    dose_end=1000,
    reps=3,
    groups=2,
    missing_data="off_axis",
) -> pl.DataFrame:
    doses = np.logspace(np.log10(dose_start), np.log10(dose_end), num=n_doses)
    doses = np.append(0, doses)
    n_doses = n_doses + 1  # To include 0
    dose_grid = [(x, y) for x in doses for y in doses]

    df = pl.DataFrame(dose_grid, orient="row")
    df = pl.concat([df] * reps * groups)
    df = df.rename({"column_0": "drug_a", "column_1": "drug_b"})

    df = df.with_columns(
        pl.Series(
            "drug_a_resp",
            [ll4(x, 1, 0, 100, 10) if x > 0 else 0.0 for x in df["drug_a"]],
        ),
        pl.Series(
            "drug_b_resp",
            [ll4(x, 1, 0, 100, 10) if x > 0 else 0.0 for x in df["drug_b"]],
        ),
        pl.Series("group", np.repeat(range(groups), n_doses**2 * reps)),
    ).with_columns((bliss_response(col("drug_a_resp"), col("drug_b_resp"))).alias("resp"))

    if missing_data == "off_axis":
        df = _remove_off_axis(df, ["drug_a", "drug_b"], "resp")

    df = df.drop("drug_a_resp", "drug_b_resp")
    return df


def make_example_resp(
    doses,
    slope=1,
    min_val=0,
    max_val=100,
    ic50=10,
    add_error=False,
):
    if add_error:
        doses += doses * np.random.normal(size=len(doses), scale=0.05)
    return pl.Series(
        [ll4(x, slope, min_val, max_val, ic50) if x > 0 else 0.0 for x in doses]
    )


def bliss_response(x, y):
    return x + y - x * y / 100


def _remove_off_axis(df, drug_cols, resp_col):
    a, b = drug_cols
    return df.with_columns(
        pl.when((col(a) == col(b)) | (col(a) == 0) | (col(b) == 0))
        .then(col(resp_col))
        .otherwise(np.nan)
    )
